
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
import os

# Load the workbook
workbook = load_workbook('/Users/marufahmed/Work/Apurba/ocr_dataset_evaluation/reporting/analysis/ccw_selection.xlsx')
sheet = workbook.active

# Path to your image folder
image_folder_path = '/Users/marufahmed/Work/Apurba/OCR DATASET/data/ocr/ccw_word'

# Insert a new column named 'Image' in the second position (column B)
sheet.insert_cols(2)
sheet.cell(row=1, column=2, value='Image')

# Get the maximum row count in the sheet
max_row = sheet.max_row

# Loop through the rows starting from the second row (assuming the first row contains headers)
for row in range(2, max_row + 1):
    # Get the Image_ID from each row
    image_id = sheet.cell(row, 1).value  # Assuming Image_ID is in column A (index 1)
    
    # Construct the image path based on the Image_ID
    image_path = os.path.join(image_folder_path, f"{image_id}")

    # Check if the image file exists
    if os.path.isfile(image_path):
        try:
            img = ExcelImage(image_path)
            img.width = img.width * 0.5  # Adjust the image width to fit inside the cell
            img.height = img.height * 0.5  # Adjust the image height to fit inside the cell
            
            # Insert the image in the 'Image' column
            sheet.add_image(img, f"B{row}")  # Insert in column B (index 2)
        except FileNotFoundError:
            print(f"Image '{image_id}' not found at '{image_path}'")
    else:
        print(f"Image '{image_id}' not found at '{image_path}'")

# Save the modified workbook
workbook.save('modified_excel_with_images.xlsx')

